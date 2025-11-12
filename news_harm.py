#!/usr/bin/env python3
"""
News Market Bot (Excel-enabled)

Outputs:
  1) news_bot_output/news_db.xlsx
     Persistent database (Date, Source, Title, Link, Sentiment, ID, Tickers, Companies)
     Appends and de-duplicates by ID. Keeps newest row on conflicts.

  2) news_bot_output/news_last7.xlsx
     Rolling last N days (Date, Source, Title, Summary, Link, Sentiment, ID, Tickers, Companies)
     With formatting: Summary ≈ 500 px, Source ≈ 220 px, and external borders per data row.

  3) news_bot_output/news_outputs.xlsx
     Legacy multi-sheet (RawNews, MappedScored, DailySignals)

Run example:
  py .\news_harm.py --tickers EXC XEL MSFT --backend vader --days 7
"""

import argparse
import datetime as dt
import hashlib
import json
import os
import re
import sys
from typing import Dict, List

import feedparser
import pandas as pd

# Sentiment: VADER (built-in via nltk)
from nltk.sentiment import SentimentIntensityAnalyzer
import nltk

# Optional: FinBERT
try:
    from transformers import AutoTokenizer, AutoModelForSequenceClassification
    import torch
    FINBERT_AVAILABLE = True
except Exception:
    FINBERT_AVAILABLE = False

# Optional: prices
try:
    import yfinance as yf
    YF_AVAILABLE = True
except Exception:
    YF_AVAILABLE = False

# Optional: plotting
try:
    import plotly.graph_objects as go
    PLOTLY_AVAILABLE = True
except Exception:
    PLOTLY_AVAILABLE = False

# Excel formatting helpers
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter

# Project-local aliases
from ticker_aliases import DEFAULT_TICKERS, build_aliases

# -----------------------------
# Config
# -----------------------------
if os.path.exists("aliases.json"):
    with open("aliases.json", "r", encoding="utf-8") as f:
        TICKER_ALIASES = json.load(f)
else:
    TICKER_ALIASES = build_aliases(DEFAULT_TICKERS)
    with open("aliases.json", "w", encoding="utf-8") as f:
        json.dump(TICKER_ALIASES, f, ensure_ascii=False, indent=2)

AR_FEEDS = [
    "https://www.ambito.com/rss/economia.xml",
    "https://www.ambito.com/rss/finanzas.xml",
    "https://www.cronista.com/rss/feed.xml",
    "https://www.infobae.com/argentina-footer/infobae/rss/",
]

GLOBAL_FEEDS = [
    "https://www.reutersagency.com/feed/?best-topics=business-finance&post_type=best",
    "https://feeds.a.dj.com/rss/RSSMarketsMain.xml",
    "https://www.cnbc.com/id/100003114/device/rss/rss.html",
    "https://www.investing.com/rss/news_25.rss",
]

GENERAL_FEEDS = GLOBAL_FEEDS + AR_FEEDS
YF_TICKER_FEED = "https://feeds.finance.yahoo.com/rss/2.0/headline?s={ticker}&region=US&lang=en-US"

DATA_DIR = "news_bot_output"
os.makedirs(DATA_DIR, exist_ok=True)

DB_XLSX = os.path.join(DATA_DIR, "news_db.xlsx")
LASTN_XLSX = os.path.join(DATA_DIR, "news_last7.xlsx")  # name kept, respects --days window

# -----------------------------
# Utils
# -----------------------------
def md5(text: str) -> str:
    return hashlib.md5(text.encode("utf-8")).hexdigest()

def normalize_text(text: str) -> str:
    if not text:
        return ""
    text = re.sub(r"http\S+", "", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text

def parse_date(entry) -> dt.date:
    for key in ("published_parsed", "updated_parsed"):
        if getattr(entry, key, None):
            tstruct = getattr(entry, key)
            try:
                return dt.date(*tstruct[:3])
            except Exception:
                pass
    return dt.date.today()

# -----------------------------
# Fetch news
# -----------------------------
def fetch_feeds(tickers: List[str]) -> pd.DataFrame:
    feeds = list(GENERAL_FEEDS) + [YF_TICKER_FEED.format(ticker=t) for t in tickers]
    rows = []
    for url in feeds:
        try:
            parsed = feedparser.parse(url)
            for e in parsed.entries:
                title = normalize_text(getattr(e, "title", ""))
                summary = normalize_text(getattr(e, "summary", ""))
                link = getattr(e, "link", "")
                date = parse_date(e)
                uid = md5((title or "") + (summary or "") + (link or ""))
                rows.append(
                    {
                        "uid": uid,
                        "date": pd.to_datetime(date),
                        "title": title,
                        "summary": summary,
                        "link": link,
                        "source": parsed.feed.get("title", url),
                    }
                )
        except Exception as ex:
            print(f"[warn] failed feed: {url} -> {ex}", file=sys.stderr)

    df = pd.DataFrame(rows).drop_duplicates(subset=["uid"])
    if df.empty:
        return df
    df["date"] = pd.to_datetime(df["date"], errors="coerce")
    df = df.sort_values("date", ascending=False).reset_index(drop=True)
    return df

# -----------------------------
# Sentiment backends
# -----------------------------
# Ensure VADER lexicon
try:
    nltk.data.find("sentiment/vader_lexicon.zip")
except LookupError:
    nltk.download("vader_lexicon")

class VaderBackend:
    def __init__(self):
        self.analyzer = SentimentIntensityAnalyzer()
    def score(self, text: str) -> float:
        if not text:
            return 0.0
        return self.analyzer.polarity_scores(text)["compound"]

class FinBERTBackend:
    def __init__(self):
        if not FINBERT_AVAILABLE:
            raise RuntimeError("FinBERT backend not available. Install transformers + torch.")
        self.tokenizer = AutoTokenizer.from_pretrained("ProsusAI/finbert")
        self.model = AutoModelForSequenceClassification.from_pretrained("ProsusAI/finbert")
        self.model.eval()
    @torch.no_grad()
    def score(self, text: str) -> float:
        if not text:
            return 0.0
        inputs = self.tokenizer(text, return_tensors="pt", truncation=True, max_length=256)
        outputs = self.model(**inputs)
        probs = torch.nn.functional.softmax(outputs.logits, dim=-1).flatten()
        neg, neu, pos = probs.tolist()
        return float(pos - neg)

def get_backend(name: str):
    if name.lower() == "vader":
        return VaderBackend()
    if name.lower() == "finbert":
        return FinBERTBackend()
    raise ValueError(f"Unknown backend: {name}")

# -----------------------------
# Mapping helpers
# -----------------------------
def build_ticker_regexes(tickers: List[str], aliases_map: Dict[str, List[str]]) -> Dict[str, re.Pattern]:
    patterns = {}
    for t in tickers:
        words = [re.escape(t)]
        for alias in aliases_map.get(t, []):
            words.append(re.escape(alias))
        pat = r"(?i)\b(" + r"|".join(words) + r")\b"
        patterns[t] = re.compile(pat, flags=re.IGNORECASE)
    return patterns

def map_articles_to_tickers(df_news: pd.DataFrame, tickers: List[str]) -> pd.DataFrame:
    regs = build_ticker_regexes(tickers, TICKER_ALIASES)
    rows = []
    for _, r in df_news.iterrows():
        text = f"{r['title']} {r['summary']}".strip()
        matched = []
        for t in tickers:
            if regs[t].search(text):
                matched.append(t)
        if not matched:
            matched = ["MARKET"]
        for t in matched:
            rows.append(
                {
                    "date": r["date"].date(),
                    "ticker": t,
                    "title": r["title"],
                    "summary": r["summary"],
                    "source": r["source"],
                    "link": r["link"],
                    "uid": r["uid"],
                }
            )
    if not rows:
        return pd.DataFrame(columns=["date", "ticker", "title", "summary", "source", "link", "uid"])
    return pd.DataFrame(rows)

def score_articles(df_mapped: pd.DataFrame, backend_name: str) -> pd.DataFrame:
    backend = get_backend(backend_name)
    scores = []
    for _, r in df_mapped.iterrows():
        text = (r["title"] or "") + ". " + (r["summary"] or "")
        s = backend.score(text)
        scores.append(s)
    df_mapped = df_mapped.copy()
    df_mapped["sentiment"] = scores
    return df_mapped

def aggregate_daily(df_scored: pd.DataFrame) -> pd.DataFrame:
    agg = (
        df_scored.groupby(["date", "ticker"])
        .agg(
            mean_sentiment=("sentiment", "mean"),
            n_articles=("uid", "nunique"),
        )
        .reset_index()
        .sort_values(["ticker", "date"])
    )
    def signal(x: float) -> str:
        if x >= 0.15:
            return "BUY"
        if x <= -0.15:
            return "SELL"
        return "HOLD"
    agg["signal"] = agg["mean_sentiment"].apply(signal)
    return agg

# -----------------------------
# Companies (names) annotation
# -----------------------------
def build_company_names(tickers: List[str]) -> Dict[str, str]:
    names: Dict[str, str] = {}
    for t in tickers:
        nm = None
        if YF_AVAILABLE:
            try:
                info = yf.Ticker(t).get_info()
                nm = info.get("longName") or info.get("shortName")
            except Exception:
                nm = None
        names[t] = nm or t
    names["MARKET"] = "Market"
    return names

def annotate_companies(
    df_news: pd.DataFrame,
    tickers: List[str],
    aliases_map: Dict[str, List[str]],
    company_names: Dict[str, str],
) -> pd.DataFrame:
    regs = build_ticker_regexes(tickers, aliases_map)
    tickers_col, companies_col = [], []
    for _, r in df_news.iterrows():
        text = f"{r.get('title','')} {r.get('summary','')}".strip()
        matched = [t for t, rgx in regs.items() if rgx.search(text)]
        matched = sorted(set(matched))
        tickers_col.append(",".join(matched))
        companies_col.append(",".join(company_names.get(t, t) for t in matched))
    out = df_news.copy()
    out["tickers"] = tickers_col
    out["companies"] = companies_col
    return out

# -----------------------------
# Prices & plotting (optional)
# -----------------------------
def add_returns(daily: pd.DataFrame, lookahead_days: int = 1) -> pd.DataFrame:
    if not YF_AVAILABLE:
        return daily
    out = []
    for tkr, d in daily.groupby("ticker"):
        y_ticker = "^GSPC" if tkr == "MARKET" else tkr
        try:
            start = (pd.to_datetime(d["date"].min()) - pd.Timedelta(days=7)).date()
            end = (pd.to_datetime(d["date"].max()) + pd.Timedelta(days=7)).date()
            px = yf.download(y_ticker, start=str(start), end=str(end), progress=False)["Adj Close"].dropna()
            df = d.copy().sort_values("date")
            df["date"] = pd.to_datetime(df["date"])
            px = px.asfreq("B").ffill()
            next_px = px.shift(-lookahead_days)
            ret = (next_px / px - 1.0).rename("fwd_return")
            joined = df.set_index("date").join(px.rename("price")).join(ret)
            joined["ticker"] = tkr
            out.append(joined.reset_index())
        except Exception as ex:
            print(f"[warn] price fetch failed for {y_ticker}: {ex}", file=sys.stderr)
    if out:
        res = pd.concat(out, ignore_index=True)
        keep = [c for c in ["date", "ticker", "mean_sentiment", "n_articles", "signal", "price", "fwd_return"] if c in res.columns]
        return res[keep]
    return daily

def plot_ticker(daily_with_ret: pd.DataFrame, ticker: str):
    if not PLOTLY_AVAILABLE:
        print("[info] Plotly not installed; skipping plot.")
        return
    d = daily_with_ret[daily_with_ret["ticker"] == ticker].dropna(subset=["mean_sentiment"])
    if d.empty:
        print(f"[info] No data to plot for {ticker}")
        return
    fig = go.Figure()
    fig.add_trace(go.Bar(x=d["date"], y=d["mean_sentiment"], name="Daily Sentiment"))
    if "fwd_return" in d.columns and d["fwd_return"].notna().any():
        fig.add_trace(go.Scatter(x=d["date"], y=d["fwd_return"], name="Next-Day Return", mode="lines+markers", yaxis="y2"))
        fig.update_layout(
            title=f"{ticker} — Sentiment vs. Next-Day Returns",
            xaxis_title="Date", yaxis_title="Sentiment (avg)",
            yaxis2=dict(title="Fwd Return", overlaying="y", side="right"),
            legend=dict(orientation="h")
        )
    else:
        fig.update_layout(
            title=f"{ticker} — Daily Sentiment",
            xaxis_title="Date", yaxis_title="Sentiment (avg)",
            legend=dict(orientation="h")
        )
    fig.show()

# -----------------------------
# Excel writers & formatting
# -----------------------------
def _pixels_to_width(pixels: int) -> float:
    """Approximate conversion from pixels to Excel column width (Calibri 11)."""
    return max(1.0, (pixels - 5) / 7.0)

def _set_col_width_by_header(ws, header_name: str, pixels: int):
    target = header_name.strip().lower()
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=col).value
        if v is not None and str(v).strip().lower() == target:
            ws.column_dimensions[get_column_letter(col)].width = _pixels_to_width(pixels)
            break

def _add_row_box_borders(ws, data_start_row: int = 2):
    """
    For each data row, draw a thin external rectangle around the row:
    - top & bottom across all cells
    - left on first column, right on last column
    """
    thin = Side(style="thin", color="000000")
    max_row, max_col = ws.max_row, ws.max_column
    if max_row < data_start_row or max_col < 1:
        return
    for r in range(data_start_row, max_row + 1):
        for c in range(1, max_col + 1):
            left  = thin if c == 1 else None
            right = thin if c == max_col else None
            top   = thin
            bottom= thin
            ws.cell(row=r, column=c).border = Border(left=left, right=right, top=top, bottom=bottom)


def _apply_date_format(ws, header_name: str = "Date"):
    """Find a header and set yyyy-mm-dd number format for that column."""
    target = header_name.strip().lower()
    col_idx = None
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v is not None and str(v).strip().lower() == target:
            col_idx = c
            break
    if not col_idx:
        return
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=col_idx).number_format = "yyyy-mm-dd"


def save_to_excel(news: pd.DataFrame, scored: pd.DataFrame, daily: pd.DataFrame) -> str:
    xlsx_path = os.path.join(DATA_DIR, "news_outputs.xlsx")

    n2, s2, d2 = news.copy(), scored.copy(), daily.copy()
    for df in (n2, s2, d2):
        if "date" in df.columns:
            df["date"] = pd.to_datetime(df["date"], errors="coerce").dt.date

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        n2.to_excel(writer, sheet_name="RawNews", index=False)
        s2.to_excel(writer, sheet_name="MappedScored", index=False)
        d2.to_excel(writer, sheet_name="DailySignals", index=False)

        for sheet in ("RawNews", "MappedScored", "DailySignals"):
            ws = writer.sheets[sheet]
            # try both casings, whichever exists
            _set_col_width_by_header(ws, "Date", 80)
            _set_col_width_by_header(ws, "date", 80)
            _apply_date_format(ws, "date")  # header might be lower-case here

    print(f"[info] Excel overwritten: {xlsx_path}")
    return xlsx_path


# --- Persistent DB + Weekly ---
def _score_raw_news(df_news: pd.DataFrame, backend_name: str) -> pd.DataFrame:
    if df_news is None or df_news.empty:
        return df_news.assign(sentiment=pd.Series(dtype=float))
    backend = get_backend(backend_name)
    txt = (df_news["title"].fillna("") + ". " + df_news["summary"].fillna("")).tolist()
    df = df_news.copy()
    df["sentiment"] = [backend.score(t) for t in txt]
    return df

def _to_db_frame(df: pd.DataFrame) -> pd.DataFrame:
    """DB schema: Date, Source, Title, Link, Sentiment, ID, Tickers, Companies."""
    base_cols = ["date", "source", "title", "link", "sentiment", "uid"]
    opt_cols = ["tickers", "companies"]
    cols = [c for c in base_cols + opt_cols if c in df.columns]
    out = df.loc[:, cols].copy()
    rename_map = {
        "date": "Date",
        "source": "Source",
        "title": "Title",
        "link": "Link",
        "sentiment": "Sentiment",
        "uid": "ID",
        "tickers": "Tickers",
        "companies": "Companies",
    }
    out = out.rename(columns={c: rename_map.get(c, c) for c in out.columns})
    out["Date"] = pd.to_datetime(out["Date"], errors="coerce").dt.date
    for must in ["Date", "Source", "Title", "Link", "Sentiment", "ID", "Tickers", "Companies"]:
        if must not in out.columns:
            out[must] = pd.NA
    return out[["Date", "Source", "Title", "Link", "Sentiment", "ID", "Tickers", "Companies"]]

def _read_existing_db(path: str) -> pd.DataFrame:
    if os.path.exists(path):
        try:
            df = pd.read_excel(path, sheet_name=0, engine="openpyxl")
            # tolerate any order/casing; the updater will reindex
            return df
        except Exception:
            pass
    return pd.DataFrame()

def _update_db_excel(db_path: str, new_rows: pd.DataFrame) -> str:
    """Append + de-dupe by ID, keep most recent Date, and format widths/borders."""
    prev = _read_existing_db(db_path)
    cols = ["Date", "Source", "Title", "Link", "Sentiment", "ID", "Tickers", "Companies"]

    # Bring new rows to DB schema
    tmp = new_rows.copy()
    tmp = tmp.rename(columns={c: c.strip().title() for c in tmp.columns})
    for c in cols:
        if c not in tmp.columns:
            tmp[c] = pd.NA
    new_rows = tmp[cols]

    all_rows = new_rows.copy() if prev.empty else pd.concat([prev, new_rows], ignore_index=True)
    if "Date" in all_rows.columns:
        all_rows["Date"] = pd.to_datetime(all_rows["Date"], errors="coerce").dt.date
    all_rows = all_rows.sort_values("Date", ascending=False).drop_duplicates(subset=["ID"], keep="first")

    with pd.ExcelWriter(db_path, engine="openpyxl") as w:
        all_rows.to_excel(w, sheet_name="News", index=False)
        ws = w.sheets["News"]
        _set_col_width_by_header(ws, "Date", 80)      # << add
        _set_col_width_by_header(ws, "Source", 220)
        _add_row_box_borders(ws, data_start_row=2)
        _apply_date_format(ws, "Date")

    print(f"[info] DB updated: {db_path} (rows={len(all_rows)})")
    return db_path

def _write_lastn_excel(path: str, df_lastn: pd.DataFrame) -> str:
    """
    Last N days with:
    Date, Source, Title, Summary, Link, Sentiment, ID, Tickers, Companies
    + formatting (Summary≈500px, Source≈220px, row borders).
    """
    cols = ["date", "source", "title", "summary", "link", "sentiment", "uid", "tickers", "companies"]
    cols = [c for c in cols if c in df_lastn.columns]
    out = df_lastn.loc[:, cols].copy()

    rename_map = {
        "date": "Date",
        "source": "Source",
        "title": "Title",
        "summary": "Summary",
        "link": "Link",
        "sentiment": "Sentiment",
        "uid": "ID",
        "tickers": "Tickers",
        "companies": "Companies",
    }
    out = out.rename(columns={c: rename_map.get(c, c) for c in out.columns})
    out["Date"] = pd.to_datetime(out["Date"], errors="coerce").dt.date
    out = out.sort_values("Date", ascending=False)

    with pd.ExcelWriter(path, engine="openpyxl") as w:
        out.to_excel(w, sheet_name="LastN", index=False)
        ws = w.sheets["LastN"]
        _set_col_width_by_header(ws, "Date", 80)      # << add
        _set_col_width_by_header(ws, "Summary", 500)
        _set_col_width_by_header(ws, "Source", 220)
        _add_row_box_borders(ws, data_start_row=2)
        _apply_date_format(ws, "Date")


    print(f"[info] Weekly written: {path} (rows={len(out)})")
    return path

# -----------------------------
# Main pipeline
# -----------------------------
def run(tickers: List[str], backend: str, days: int, plot: bool, lookahead: int):
    print(f"[info] tickers={tickers} backend={backend} days={days}")

    # 1) Fetch all items
    news_all = fetch_feeds(tickers)
    if news_all.empty:
        print("[warn] no news found")
        return

    # 2) Annotate tickers/companies
    company_names = build_company_names(tickers)
    news_all = annotate_companies(news_all, tickers, TICKER_ALIASES, company_names)

    # 3) Score sentiment once per unique article
    raw_scored_all = _score_raw_news(news_all, backend)

    # 4) Update persistent DB
    db_frame = _to_db_frame(raw_scored_all)
    _update_db_excel(DB_XLSX, db_frame)

    # 5) Last N days snapshot
    cutoff = pd.Timestamp.today().normalize() - pd.Timedelta(days=days)
    lastn = raw_scored_all[raw_scored_all["date"] >= cutoff].copy()
    _write_lastn_excel(LASTN_XLSX, lastn)

    # 6) Legacy pipeline (mapped/scored/aggregated) over last N
    mapped = map_articles_to_tickers(lastn, tickers)
    scored = score_articles(mapped, backend)
    daily = aggregate_daily(scored)
    daily = add_returns(daily, lookahead_days=lookahead)
    xlsx_path = save_to_excel(lastn, scored, daily)

    # 7) Console signals
    print("\n=== Signals (last {} days) ===".format(days))
    if not daily.empty:
        latest_day = daily["date"].max()
        latest = daily[pd.to_datetime(daily["date"]) == pd.to_datetime(latest_day)]
        latest = latest.sort_values(["ticker"])
        for _, r in latest.iterrows():
            sent = float(r["mean_sentiment"])
            n = int(r["n_articles"])
            sig = r["signal"]
            print(f"{r['ticker']:<6}  signal={sig:<4}  sentiment={sent:+.3f}  n={n}")

    print("\nSaved:")
    print(f"- DB          : {DB_XLSX}")
    print(f"- Last {days}d: {LASTN_XLSX}")
    print(f"- Legacy XLSX : {xlsx_path}")

    plot = plot or (os.getenv("PLOTLY_ENABLED", "0") == "1")
    if plot:
        print("\n[info] Generating plots...")
        for t in tickers:
            plot_ticker(daily, t)
    else:
        print("[info] Plotting skipped (use --plot to enable)")

def parse_args():
    p = argparse.ArgumentParser(description="News Market Bot (Excel-enabled)")
    p.add_argument("--tickers", nargs="+", default=DEFAULT_TICKERS, help="List of tickers")
    p.add_argument("--backend", default="vader", choices=["vader", "finbert"], help="Sentiment backend")
    p.add_argument("--days", type=int, default=7, help="Lookback window for news")
    p.add_argument("--plot", action="store_true", help="Plot sentiment vs returns")
    p.add_argument("--lookahead", type=int, default=1, help="Days ahead to compute forward return")
    args, _ = p.parse_known_args()
    return args

if __name__ == "__main__":
    args = parse_args()
    try:
        run(args.tickers, args.backend, args.days, args.plot, args.lookahead)
    except KeyboardInterrupt:
        print("\nInterrupted by user")

 